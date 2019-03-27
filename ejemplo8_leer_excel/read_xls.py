from openpyxl import *


def read_excel_data(filename):
    data = dict()

    content = load_workbook(filename)

    sheet = content.active

    abs = []

    for j in range(1,15):
        i = 1
        while i < 5 and sheet.cell(row=i,column=j).value == None:
            i += 1
        if sheet.cell(row=i,column=j).value == None:
            continue

        name = sheet.cell(row=i,column=j).value

        col_data = []
        i += 1
        while sheet.cell(row=i,column=j).value != None:
            col_data.append(sheet.cell(row=i,column=j).value)
            i += 1
        data[name ] = col_data

    return data

#data = read_excel_data("Mediciones/Ej4_Bode_k0.xlsx")
#print(data)
